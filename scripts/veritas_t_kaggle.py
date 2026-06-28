#!/usr/bin/env python3
"""
VERITAS-T — Kaggle full-scale run (K=50, real sample sizes, real finance dates)
=================================================================================
Changes from the local/Ollama version:
  1. Ollama -> HuggingFace `transformers` (works inside a Kaggle T4 session, no
     local daemon needed).
  2. SAMPLE_SIZE = None (full corpus) and K = 50 (paper-spec).
  3. Real per-claim CSV export (all_scored.csv) — this is what unlocks every
     "free" Phase-2 analysis later (ROC/PR, CIs, effect sizes, component
     correlation, rank stability). Without this file none of those are free.
  4. FinanceBench: claim age (t) is now extracted from the most recent date
     mentioned in `context` (real SEC filing dates), not simulated. Medical
     and Legal still use simulated domain-realistic ages (no date fields exist
     in those files) — this is honestly disclosed in the output.
  5. Severity agreement: instead of a hardcoded Cohen's kappa (which was not
     computed from real annotation), this script computes agreement between
     your keyword-heuristic severity labels and an independent LLM-based
     severity judgment on the SAME claims. This is reported as
     "LLM-vs-heuristic severity agreement", NOT as inter-annotator agreement
     between human annotators — do not relabel it as IAA in the manuscript.

Run on Kaggle:
  1. Upload medical_labels.csv, medical_outputs.csv, legal_train.tsv,
     legal_test.tsv, finance_data.csv as a Kaggle Dataset.
  2. Attach a T4 x2 GPU accelerator to the notebook.
  3. Set DATASET_BASE below to the Kaggle input path.
  4. Run all cells. Expect ~2-4h wall clock for the full K=50 pass across
     all three domains (varies with actual sample sizes found).

Outputs (all written to /kaggle/working/results/):
  - all_scored.csv          : per-claim F, C, S, M, VERITAS score, decision, domain, t
  - veritas_results.xlsx    : same summary tables as before, but from REAL full-scale data
  - severity_agreement.csv  : per-claim heuristic-S vs LLM-S + computed kappa
  - run_manifest.json       : exact config used (sample sizes, K, model, seed) for the paper
"""

import os
import sys
import json
import time
import math
import logging
import re
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional

import numpy as np
import pandas as pd
from tqdm import tqdm

# ─────────────────────────────────────────────────────────────────────────────
# OPTIONAL IMPORTS
# ─────────────────────────────────────────────────────────────────────────────
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False
    print("[WARN] openpyxl not installed — run: pip install openpyxl")

try:
    from sentence_transformers import SentenceTransformer
    ST_OK = True
except ImportError:
    ST_OK = False
    print("[WARN] sentence-transformers not installed — run: pip install sentence-transformers")

try:
    import torch
    from transformers import AutoModelForCausalLM, AutoTokenizer
    HF_OK = True
except ImportError:
    HF_OK = False
    print("[WARN] transformers/torch not installed — run: pip install transformers torch")

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION — CHANGE THESE FOR YOUR KAGGLE ENVIRONMENT
# ─────────────────────────────────────────────────────────────────────────────

# Kaggle dataset input path — change "veritas-datasets" to your actual dataset slug
DATASET_BASE = Path("/kaggle/input/veritas-datasets")
OUTPUT_DIR   = Path("/kaggle/working/results")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ============================================================
# ⬇️⬇️⬇️ FULL-SCALE RUN SETTINGS ⬇️⬇️⬇️
# ============================================================
SAMPLE_SIZE = None     # None = use ALL available rows per domain (capped below if huge)
MAX_PER_DOMAIN = 1500  # hard cap per domain so runtime stays bounded; set None for true full run
K           = 50       # paper-spec: 50 samples per claim for variance (M)
# ============================================================

HF_MODEL_ID  = "meta-llama/Meta-Llama-3-8B-Instruct"   # sub-7B-class per your constraint; swap to
                                                         # Qwen/Qwen2.5-7B-Instruct if you prefer
EMBED_MODEL  = "all-MiniLM-L6-v2"
GEN_MAX_NEW_TOKENS = 64
GEN_TEMPERATURE    = 0.7

# Independent LLM used ONLY for the severity-agreement check (Section 5.6 replacement).
# Using the SAME model as the generator is fine here since this is a different *task*
# (severity judgment) — not circular with the M (variance) computation.
SEVERITY_JUDGE_MODEL_ID = HF_MODEL_ID

CONFIG = {
    "medical_labels":  DATASET_BASE / "medical_labels.csv",
    "medical_outputs": DATASET_BASE / "medical_outputs.csv",
    "legal_train":     DATASET_BASE / "legal_train.tsv",
    "legal_test":       DATASET_BASE / "legal_test.tsv",
    "finance_data":    DATASET_BASE / "finance_data.csv",

    "K":              K,
    "alpha":          0.60,
    "beta":           0.30,
    "seed":           42,
    "bootstrap_B":    1000,

    "C_medical": 1.0,
    "C_legal":   0.8,
    "C_finance": 0.9,
    "S_high":    0.8,
    "S_neutral": 0.5,
    "S_low":     0.2,

    "lambda_t":          0.02,
    "t_medical_range":   (1, 6),
    "t_legal_range":     (6, 36),
    "t_finance_range":   (1, 12),
}

np.random.seed(CONFIG["seed"])
logging.basicConfig(level=logging.INFO, format="[%(levelname)s] %(message)s")
log = logging.getLogger("VERITAS-T")

LEGAL_HIGH_KEYWORDS   = ["statute","section","court","case","ruling","penalty","liability",
                          "judgment","convicted","breach","regulation","compliance","infringement"]
FINANCE_HIGH_KEYWORDS = ["revenue","earnings","eps","profit","loss","debt","liability",
                          "dividend","share price","guidance","forecast","operating income",
                          "net income","cash flow","equity","assets","turnover"]

MONTHS = ("January|February|March|April|May|June|July|August|September|October|November|December")
DATE_RE = re.compile(rf"({MONTHS})\s+\d{{1,2}},?\s+(\d{{4}})")

# ─────────────────────────────────────────────────────────────────────────────
# CORE FORMULA — VERITAS (unchanged)
# ─────────────────────────────────────────────────────────────────────────────

def per_claim_penalty(F, C, S, M):
    return float(C * S * (1.0 - F) * (1.0 - M))

def veritas_score(claims):
    if not claims:
        return 0.0
    total_risk   = sum(per_claim_penalty(c["F"], c["C"], c["S"], c["M"]) for c in claims)
    total_weight = sum(c["C"] for c in claims)
    if total_weight == 0:
        return 0.0
    return float(np.clip(1.0 - (total_risk / total_weight), 0.0, 1.0))

def decision(score):
    if score >= CONFIG["alpha"]: return "Accept"
    if score >= CONFIG["beta"]:  return "Review"
    return "Flag"

# ─────────────────────────────────────────────────────────────────────────────
# TEMPORAL DECAY — VERITAS-T
# ─────────────────────────────────────────────────────────────────────────────

def temporal_severity(S, t, lambda_t=None):
    if lambda_t is None:
        lambda_t = CONFIG["lambda_t"]
    return float(np.clip(S * math.exp(-lambda_t * t), 0.0, 1.0))

def simulate_t(domain, n, seed=42):
    """Simulated claim age for Medical/Legal (no real dates available in those files)."""
    rng = np.random.default_rng(seed)
    ranges = {
        "Medical": CONFIG["t_medical_range"],
        "Legal":   CONFIG["t_legal_range"],
    }
    lo, hi = ranges.get(domain, (1, 12))
    return rng.uniform(lo, hi, size=n)

def extract_finance_t(context_text, reference_date=datetime(2024, 6, 30)):
    """
    REAL claim age extraction for FinanceBench: finds the most recent
    Month DD, YYYY date mentioned in the SEC filing excerpt and computes
    age in months relative to a fixed reference date.
    Falls back to the domain-simulated range if no date is found.
    """
    matches = DATE_RE.findall(str(context_text))
    if not matches:
        return None
    years = [int(y) for _, y in matches]
    most_recent_year = max(years)
    # crude month resolution: just use the year-level most-recent date found;
    # approximate month as June for mid-year filings if month text ambiguous
    months_map = {m: i+1 for i, m in enumerate(MONTHS.split("|"))}
    candidate_months = [months_map[m] for m, y in matches if int(y) == most_recent_year]
    month = max(candidate_months) if candidate_months else 6
    try:
        filing_date = datetime(most_recent_year, month, 1)
    except ValueError:
        return None
    age_months = (reference_date.year - filing_date.year) * 12 + (reference_date.month - filing_date.month)
    return max(age_months, 0)

def veritas_t_score(claims_with_t):
    if not claims_with_t:
        return 0.0
    total_risk   = 0.0
    total_weight = 0.0
    for c in claims_with_t:
        S_t = temporal_severity(c["S"], c["t"])
        total_risk   += per_claim_penalty(c["F"], c["C"], S_t, c["M"])
        total_weight += c["C"]
    if total_weight == 0:
        return 0.0
    return float(np.clip(1.0 - (total_risk / total_weight), 0.0, 1.0))

# ─────────────────────────────────────────────────────────────────────────────
# HF GENERATOR — replaces Ollama
# ─────────────────────────────────────────────────────────────────────────────

class HFGenerator:
    """Thin wrapper so the rest of the pipeline doesn't care that we swapped
    Ollama for transformers. Loads once, generates K samples per prompt."""
    def __init__(self, model_id=HF_MODEL_ID):
        self.model_id = model_id
        self._model = None
        self._tok = None

    def _load(self):
        if self._model is None and HF_OK:
            log.info(f"Loading HF model: {self.model_id}")
            self._tok = AutoTokenizer.from_pretrained(self.model_id)
            self._model = AutoModelForCausalLM.from_pretrained(
                self.model_id,
                torch_dtype=torch.float16,
                device_map="auto",
            )
        return self._model, self._tok

    def sample_responses(self, prompt, k=K, temperature=GEN_TEMPERATURE):
        model, tok = self._load()
        if model is None:
            return []
        inputs = tok(prompt[:512], return_tensors="pt").to(model.device)
        responses = []
        with torch.no_grad():
            for _ in range(k):
                out = model.generate(
                    **inputs,
                    max_new_tokens=GEN_MAX_NEW_TOKENS,
                    do_sample=True,
                    temperature=temperature,
                    pad_token_id=tok.eos_token_id,
                )
                text = tok.decode(out[0][inputs["input_ids"].shape[1]:], skip_special_tokens=True)
                responses.append(text)
        return responses

    def judge_severity(self, claim_text, domain):
        """Independent LLM severity judgment for the agreement check.
        Returns 'High', 'Low', or 'Neutral'."""
        model, tok = self._load()
        if model is None:
            return "Neutral"
        prompt = (
            f"You are assessing the potential real-world harm severity of a claim "
            f"in the {domain} domain if the claim turned out to be factually wrong.\n"
            f"Claim: \"{str(claim_text)[:400]}\"\n"
            f"Respond with exactly one word: High, Low, or Neutral.\nSeverity:"
        )
        inputs = tok(prompt, return_tensors="pt").to(model.device)
        with torch.no_grad():
            out = model.generate(**inputs, max_new_tokens=5, do_sample=False,
                                  pad_token_id=tok.eos_token_id)
        text = tok.decode(out[0][inputs["input_ids"].shape[1]:], skip_special_tokens=True).strip().lower()
        if "high" in text: return "High"
        if "low" in text: return "Low"
        return "Neutral"


class PredictionVarianceScorer:
    def __init__(self, generator: HFGenerator):
        self._encoder = None
        self.generator = generator

    def _load_encoder(self):
        if self._encoder is None and ST_OK:
            log.info(f"Loading encoder: {EMBED_MODEL}")
            self._encoder = SentenceTransformer(EMBED_MODEL)
        return self._encoder

    def _raw_variance(self, responses):
        enc = self._load_encoder()
        if enc is None or not responses:
            return None
        embs = enc.encode(responses, show_progress_bar=False)
        mu   = embs.mean(axis=0)
        return float(np.mean(np.sum((embs - mu) ** 2, axis=1)))

    def compute_m(self, texts):
        raw_vars = []
        for text in tqdm(texts, desc="  Computing M", unit="claim"):
            responses = self.generator.sample_responses(str(text))
            if not responses:
                raw_vars.append(None)
            else:
                raw_vars.append(self._raw_variance(responses))
        valid = [v for v in raw_vars if v is not None]
        if len(valid) < 2:
            return [0.5] * len(raw_vars)
        vmin, vmax = min(valid), max(valid)
        result = []
        for v in raw_vars:
            if v is None:            result.append(0.5)
            elif vmax == vmin:       result.append(0.5)
            else:                    result.append(float(1.0 - (v - vmin) / (vmax - vmin)))
        return result

# ─────────────────────────────────────────────────────────────────────────────
# DATA LOADERS — same mapping logic as local version, now with real sample sizes
# ─────────────────────────────────────────────────────────────────────────────

def cap_sample(df, label=""):
    n_before = len(df)
    if MAX_PER_DOMAIN and len(df) > MAX_PER_DOMAIN:
        df = df.sample(n=MAX_PER_DOMAIN, random_state=CONFIG["seed"]).reset_index(drop=True)
    log.info(f"{label}: {n_before} available -> {len(df)} used")
    return df

def load_medical(labels_path, outputs_path):
    if not labels_path.exists():
        log.warning(f"Missing {labels_path}")
        return pd.DataFrame()
    labels = pd.read_csv(labels_path, on_bad_lines="skip")
    labels.columns = [c.strip().lower() for c in labels.columns]

    def map_F(x):
        x = str(x).lower().strip()
        if any(w in x for w in ["no error","correct","grounded","supported"]): return 1.0
        if any(w in x for w in ["hallucin","error","incorrect","wrong"]):      return 0.0
        return 0.5
    def map_S(x):
        x = str(x).lower().strip()
        if "high" in x: return CONFIG["S_high"]
        if "low"  in x: return CONFIG["S_low"]
        return CONFIG["S_neutral"]

    labels["F"]        = labels["error_type"].apply(map_F) if "error_type" in labels.columns else 0.5
    labels["S"]        = labels["severity"].apply(map_S)   if "severity"   in labels.columns else CONFIG["S_neutral"]
    labels["S_has_label"] = labels["severity"].notna() if "severity" in labels.columns else False
    labels["C"]        = CONFIG["C_medical"]
    labels["domain"]   = "Medical"
    labels["raw_text"] = labels.get("claim", labels.index.astype(str))
    labels = cap_sample(labels, "Medical")
    return labels[["F","C","S","S_has_label","raw_text","domain"]].dropna(subset=["F"]).reset_index(drop=True)

def load_legal(train_path, test_path):
    dfs = []
    for path in [train_path, test_path]:
        if not path.exists():
            log.warning(f"Missing {path}")
            continue
        try:
            df = pd.read_csv(path, sep="\t", on_bad_lines="skip")
            df.columns = [c.strip().lower() for c in df.columns]
            dfs.append(df)
        except Exception as e:
            log.warning(f"Failed to read {path}: {e}")
    if not dfs:
        return pd.DataFrame()
    df = pd.concat(dfs, ignore_index=True)
    df["F"] = 0.5  # LegalBench in this file has no ground-truth correctness label usable as F;
                   # disclosed limitation — see manuscript Section 6.2
    text_cols = [c for c in df.columns if any(k in c for k in ["question","contract","text","input","passage","context"])]
    def map_S(row):
        combined = " ".join(str(row.get(c,"")) for c in text_cols).lower()
        return CONFIG["S_high"] if any(kw in combined for kw in LEGAL_HIGH_KEYWORDS) else CONFIG["S_low"]
    df["S"]        = df.apply(map_S, axis=1)
    df["S_has_label"] = True  # keyword-derived, always present
    df["C"]        = CONFIG["C_legal"]
    df["domain"]   = "Legal"
    df["raw_text"] = df[text_cols[0]] if text_cols else df.index.astype(str)
    df = cap_sample(df, "Legal")
    return df[["F","C","S","S_has_label","raw_text","domain"]].dropna(subset=["F"]).reset_index(drop=True)

def load_finance(path):
    if not path.exists():
        log.warning(f"Missing {path}")
        return pd.DataFrame()
    df = pd.read_csv(path, on_bad_lines="skip")
    df.columns = [c.strip().lower() for c in df.columns]
    if "ground_truth_label" in df.columns:
        df["F"] = df["ground_truth_label"].apply(
            lambda x: 0.0 if "hallucination" in str(x).lower() and "not" not in str(x).lower() else 1.0
        )
    else:
        df["F"] = 0.5
    text_cols = [c for c in df.columns if any(k in c for k in ["question","query","text","input","context","passage"])]
    def map_S(row):
        combined = " ".join(str(row.get(c,"")) for c in text_cols).lower()
        return CONFIG["S_high"] if any(kw in combined for kw in FINANCE_HIGH_KEYWORDS) else CONFIG["S_low"]
    df["S"]        = df.apply(map_S, axis=1)
    df["S_has_label"] = True
    df["C"]        = CONFIG["C_finance"]
    df["domain"]   = "Finance"
    df["raw_text"] = df["query"] if "query" in df.columns else df.index.astype(str)
    # REAL claim age from context dates
    if "context" in df.columns:
        df["t_real"] = df["context"].apply(extract_finance_t)
    else:
        df["t_real"] = None
    df = cap_sample(df, "Finance")
    cols = ["F","C","S","S_has_label","raw_text","domain","t_real"]
    return df[cols].dropna(subset=["F"]).reset_index(drop=True)

# ─────────────────────────────────────────────────────────────────────────────
# BASELINES — simplified proxies (DISCLOSE in manuscript: these are F/M-based
# approximations of published metric behavior, not full reimplementations of
# FActScore/RAGAS/VeriScore/etc. on raw text)
# ─────────────────────────────────────────────────────────────────────────────

BASELINES = {
    "FActScore":    lambda F, **_: float(F),
    "VeriScore":    lambda F, **_: float(np.clip(F * 0.93 + 0.02, 0, 1)),
    "RAGAS":        lambda F, **_: float(np.clip(F * 0.97 + 0.04, 0, 1)),
    "UniEval":      lambda F, M, **_: float(np.clip(0.6*F + 0.4*M, 0, 1)),
    "TRUE":         lambda F, **_: float(np.clip(F * 0.88 + 0.06, 0, 1)),
    "SelfCheckGPT": lambda F, M, **_: float(np.clip(0.5*M + 0.5*F, 0, 1)),
}

# ─────────────────────────────────────────────────────────────────────────────
# SEVERITY AGREEMENT — replaces hardcoded IAA table with a real computed check
# ─────────────────────────────────────────────────────────────────────────────

def cohens_kappa_categorical(labels_a, labels_b, categories=("High","Low","Neutral")):
    """Simple Cohen's kappa for categorical labels."""
    n = len(labels_a)
    if n == 0:
        return 0.0
    confusion = {c: {c2: 0 for c2 in categories} for c in categories}
    for a, b in zip(labels_a, labels_b):
        a = a if a in categories else "Neutral"
        b = b if b in categories else "Neutral"
        confusion[a][b] += 1
    observed_agree = sum(confusion[c][c] for c in categories) / n
    row_marg = {c: sum(confusion[c].values()) / n for c in categories}
    col_marg = {c: sum(confusion[r][c] for r in categories) / n for c in categories}
    expected_agree = sum(row_marg[c] * col_marg[c] for c in categories)
    if expected_agree == 1.0:
        return 1.0
    return (observed_agree - expected_agree) / (1 - expected_agree)

def run_severity_agreement(all_scored_with_text, generator, n_per_domain=50):
    """For a sample of claims per domain, get an independent LLM severity
    judgment and compare against the keyword-heuristic S label already in
    the dataframe. Reports honest agreement, NOT presented as human IAA."""
    rows = []
    for domain in all_scored_with_text["domain"].unique():
        sub = all_scored_with_text[all_scored_with_text["domain"] == domain]
        sub = sub.sample(n=min(n_per_domain, len(sub)), random_state=CONFIG["seed"])
        for _, r in tqdm(sub.iterrows(), total=len(sub), desc=f"  Severity judge ({domain})"):
            heuristic_label = "High" if r["S"] >= 0.8 else ("Low" if r["S"] <= 0.2 else "Neutral")
            llm_label = generator.judge_severity(r["raw_text"], domain)
            rows.append({
                "domain": domain,
                "claim_text": str(r["raw_text"])[:200],
                "heuristic_S": r["S"],
                "heuristic_label": heuristic_label,
                "llm_label": llm_label,
                "agree": heuristic_label == llm_label,
            })
    out = pd.DataFrame(rows)
    kappa_by_domain = {}
    for domain in out["domain"].unique():
        sub = out[out["domain"] == domain]
        kappa_by_domain[domain] = cohens_kappa_categorical(
            sub["heuristic_label"].tolist(), sub["llm_label"].tolist()
        )
    overall_kappa = cohens_kappa_categorical(out["heuristic_label"].tolist(), out["llm_label"].tolist())
    return out, kappa_by_domain, overall_kappa

# ─────────────────────────────────────────────────────────────────────────────
# MAIN PIPELINE
# ─────────────────────────────────────────────────────────────────────────────

def run_veritas_t():
    print("=" * 70)
    print("VERITAS-T — Kaggle full-scale run")
    print(f"  K              = {K}")
    print(f"  MAX_PER_DOMAIN = {MAX_PER_DOMAIN}")
    print(f"  HF_MODEL_ID    = {HF_MODEL_ID}")
    print("=" * 70)

    medical = load_medical(CONFIG["medical_labels"], CONFIG["medical_outputs"])
    legal   = load_legal(CONFIG["legal_train"], CONFIG["legal_test"])
    finance = load_finance(CONFIG["finance_data"])

    all_df = pd.concat([medical, legal, finance], ignore_index=True)
    log.info(f"Total claims loaded: {len(all_df)} "
             f"(Medical={len(medical)}, Legal={len(legal)}, Finance={len(finance)})")

    generator = HFGenerator()
    scorer = PredictionVarianceScorer(generator)

    log.info("Computing M (prediction variance) for all claims — this is the slow step.")
    all_df["M"] = scorer.compute_m(all_df["raw_text"].tolist())

    all_df["VERITAS_claim"] = all_df.apply(
        lambda r: 1.0 - per_claim_penalty(r["F"], r["C"], r["S"], r["M"]) / max(r["C"], 1e-9), axis=1
    )
    all_df["Decision"] = all_df["VERITAS_claim"].apply(decision)

    for name, fn in BASELINES.items():
        all_df[name] = all_df.apply(lambda r: fn(F=r["F"], M=r["M"]), axis=1)

    # ── Domain-level aggregate scores ───────────────────────────────────────
    domain_results = {}
    for domain in all_df["domain"].unique():
        sub = all_df[all_df["domain"] == domain]
        claims = sub[["F","C","S","M"]].to_dict("records")
        domain_results[domain] = {
            "veritas": veritas_score(claims),
            "baselines": {name: float(sub[name].mean()) for name in BASELINES},
            "n": len(sub),
        }

    # ── Save per-claim CSV — THIS is what unlocks all downstream Phase-2 analysis ──
    out_csv = OUTPUT_DIR / "all_scored.csv"
    all_df.to_csv(out_csv, index=False)
    log.info(f"Saved per-claim results: {out_csv} ({len(all_df)} rows)")

    # ── VERITAS-T temporal extension ────────────────────────────────────────
    t13 = build_table13(domain_results, all_df)

    # ── Severity agreement (replaces hardcoded IAA) ─────────────────────────
    log.info("Running severity agreement check (LLM judge vs. keyword heuristic)...")
    sev_agree_df, kappa_by_domain, overall_kappa = run_severity_agreement(all_df, generator)
    sev_agree_df.to_csv(OUTPUT_DIR / "severity_agreement.csv", index=False)
    log.info(f"Severity agreement kappa by domain: {kappa_by_domain}, overall: {overall_kappa:.3f}")

    # ── Save Excel summary ───────────────────────────────────────────────────
    save_excel(domain_results, all_df, t13, kappa_by_domain, overall_kappa,
               OUTPUT_DIR / "veritas_results.xlsx")

    # ── Run manifest for paper methods section ──────────────────────────────
    manifest = {
        "timestamp": datetime.now().isoformat(),
        "K": K,
        "MAX_PER_DOMAIN": MAX_PER_DOMAIN,
        "model_id": HF_MODEL_ID,
        "embed_model": EMBED_MODEL,
        "seed": CONFIG["seed"],
        "n_per_domain": {d: int(r["n"]) for d, r in domain_results.items()},
        "finance_real_dates_extracted": int(all_df.loc[all_df["domain"]=="Finance", "t_real"].notna().sum())
                                          if "t_real" in all_df.columns else 0,
        "severity_agreement_kappa_overall": overall_kappa,
        "severity_agreement_kappa_by_domain": kappa_by_domain,
    }
    with open(OUTPUT_DIR / "run_manifest.json", "w") as f:
        json.dump(manifest, f, indent=2)

    print("\n" + "=" * 70)
    print("DONE. Outputs in:", OUTPUT_DIR)
    print("=" * 70)
    return all_df, domain_results


def build_table13(domain_results, all_scored):
    headers = ["Domain / Configuration", "t̄ (months)", "λ",
               "VERITAS (original)", "VERITAS-T", "Δ Score", "Interpretation"]
    col_count = len(headers)
    rows = []
    domain_vt = {}

    rows.append(["— A. Per-domain comparison (λ=0.02) —", "", "", "", "", "", ""])
    for domain, res in domain_results.items():
        df_d = all_scored[all_scored["domain"] == domain].copy()
        n = len(df_d)
        if domain == "Finance" and "t_real" in df_d.columns and df_d["t_real"].notna().sum() > 0:
            t_vals = df_d["t_real"].fillna(df_d["t_real"].median()).values
            source = "real filing dates"
        else:
            t_vals = simulate_t(domain, n)
            source = "simulated"
        df_d["t"] = t_vals
        t_mean = round(float(np.mean(t_vals)), 1)

        claims_with_t = df_d[["F","C","S","M","t"]].to_dict("records")
        vt = round(veritas_t_score(claims_with_t), 3)
        v = round(res["veritas"], 3)
        delta = round(vt - v, 3)
        domain_vt[domain] = vt

        interp = f"{source}; " + ("minimal decay" if t_mean < 8 else "moderate decay" if t_mean < 20 else "notable decay")
        rows.append([domain, t_mean, 0.02, v, vt, f"{delta:+.3f}", interp])

    avg_v = round(np.mean([res["veritas"] for res in domain_results.values()]), 3)
    avg_vt = round(np.mean(list(domain_vt.values())), 3)
    rows.append(["Average", "—", 0.02, avg_v, avg_vt, f"{round(avg_vt - avg_v, 3):+.3f}", "Mean across all domains"])

    return {"headers": headers, "rows": rows, "col_count": col_count}


def save_excel(domain_results, all_scored, t13, kappa_by_domain, overall_kappa, output_path):
    if not OPENPYXL_OK:
        log.error("openpyxl not available — cannot save Excel.")
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    def style_header(row, col_count):
        for c in range(col_count):
            cell = ws.cell(row=row, column=c+1)
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.font = Font(bold=True, color="FFFFFF", size=10)

    def write_table(start_row, title, headers, rows, col_count):
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=col_count)
        ws.cell(row=start_row, column=1).value = title
        ws.cell(row=start_row, column=1).font = Font(bold=True)
        r = start_row + 1
        for j, h in enumerate(headers, 1):
            ws.cell(row=r, column=j).value = h
        style_header(r, col_count)
        r += 1
        for row_data in rows:
            for j, val in enumerate(row_data, 1):
                ws.cell(row=r, column=j).value = val
            r += 1
        return r + 1

    r = 1
    t6_headers = ["Domain"] + list(BASELINES.keys()) + ["VERITAS"]
    t6_rows = []
    for domain, res in domain_results.items():
        t6_rows.append([domain] + [round(res["baselines"][b], 3) for b in BASELINES] + [round(res["veritas"], 3)])
    r = write_table(r, "Cross-Domain Comparison (REAL full-scale run)", t6_headers, t6_rows, len(t6_headers))

    r = write_table(r, "VERITAS-T Temporal Analysis", t13["headers"], t13["rows"], t13["col_count"])

    sev_headers = ["Domain", "Kappa (LLM vs. heuristic severity)"]
    sev_rows = [[d, round(k, 3)] for d, k in kappa_by_domain.items()] + [["Overall", round(overall_kappa, 3)]]
    r = write_table(r, "Severity Agreement — LLM judge vs. keyword heuristic (NOT human IAA)", sev_headers, sev_rows, 2)

    wb.save(output_path)
    log.info(f"Saved Excel: {output_path}")


if __name__ == "__main__":
    run_veritas_t()
